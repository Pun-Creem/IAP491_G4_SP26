#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Eval from Pre-computed Top-K Similarity
========================================
Similar to auto_precision&recall.py but uses pre-computed top-k similarity
files instead of running the full pipeline.

Workflow:
  1. User selects sample folder (JSON reports)
  2. User selects top-k similarity folder (CSV/XLSX files named by sha256)
  3. For each sample:
     a. Extract sha256 & TTPs from JSON report
     b. Find matching similarity file in the chosen folder
     c. Read top-k neighbors (sorted by similarity desc)
     d. Assign neighbor actions from actionPerReport.xlsx
     e. Map sample TTPs -> D3FEND actions (px)
     f. Compute score(a) = W * px(a) + beta * sum(sim_j * yj(a))
  4. Calculate Precision / Recall / F1 at top-K = 1..5
  5. Save results

Config: config_eval.json (only W, beta, and data paths)

Usage:
  python eval_from_topk.py
  python eval_from_topk.py --sample-dir path/to/jsons --topk-dir path/to/sim_folder
"""

import argparse
import csv
import json
import os
import re
import sys
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import openpyxl
import pandas as pd

# ── Logging ──────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from log_utils import setup_logging

# ── Imports from auto_sim2action ──
from auto_sim2action import (
    extract_hash,
    extract_ttps,
    load_feature_list,
)

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_FILE = SCRIPT_DIR / "config_eval.json"
GROUND_TRUTH_PATH = SCRIPT_DIR / "precision&recall" / "sample_ground_truth" / "sample_ground_truth.xlsx"
RESULTS_DIR = SCRIPT_DIR / "precision&recall"
RESULTS_BASE = "precision_recall_f1_results_topk"


# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

def load_config() -> Dict[str, Any]:
    if not CONFIG_FILE.exists():
        return {}
    with open(CONFIG_FILE, encoding="utf-8") as f:
        data = json.load(f)
    cfg: Dict[str, Any] = {k: v for k, v in data.items() if not k.startswith("_")}
    for key, val in cfg.pop("paths", {}).items():
        cfg[key] = val
    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# GUI Pickers
# ─────────────────────────────────────────────────────────────────────────────

def pick_directory(title: str) -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return folder


# ─────────────────────────────────────────────────────────────────────────────
# Read pre-computed similarity file (CSV or XLSX)
# ─────────────────────────────────────────────────────────────────────────────

def read_similarity_file(fpath: Path) -> List[Tuple[str, float]]:
    """
    Read a similarity file (CSV or XLSX).
    Returns list of (neighbor_sha256, similarity_score) sorted desc by score.
    Handles column names: 'similarity_score' or 'similarity'.
    """
    neighbors = []

    if fpath.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return []
        header = [str(h).strip().lower() for h in rows[0]]
        sha_idx = 0
        sim_idx = 1
        for i, h in enumerate(header):
            if h in ("similarity_score", "similarity"):
                sim_idx = i
            if h == "sha256":
                sha_idx = i
        for row in rows[1:]:
            if row[sha_idx] is None:
                continue
            sha = str(row[sha_idx]).strip()
            sim = float(row[sim_idx])
            neighbors.append((sha, sim))
    else:
        with open(fpath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                sha = row.get("sha256", "").strip()
                sim_str = row.get("similarity_score") or row.get("similarity") or "0"
                sim = float(sim_str)
                if sha:
                    neighbors.append((sha, sim))

    neighbors.sort(key=lambda x: x[1], reverse=True)
    return neighbors


# ─────────────────────────────────────────────────────────────────────────────
# Load actionPerReport.xlsx -> hash -> set of actions
# ─────────────────────────────────────────────────────────────────────────────

def load_action_report(path: Path) -> Dict[str, Set[str]]:
    df = pd.read_excel(path)
    hash_to_acts: Dict[str, Set[str]] = {}
    for _, row in df.iterrows():
        h = str(row.get("Hash256", "")).strip()
        raw = str(row.get("Action", "")) if pd.notna(row.get("Action")) else ""
        hash_to_acts[h] = {a.strip() for a in raw.split("\n") if a.strip()}
    return hash_to_acts


# ─────────────────────────────────────────────────────────────────────────────
# Map sample TTPs -> D3FEND actions (binary)  [equivalent to Stage 6]
# ─────────────────────────────────────────────────────────────────────────────

def map_sample_actions(sample_ttps: Set[str], action_per_ttps_path: Path) -> Dict[str, int]:
    """
    Given a set of TTPs for a sample, map them to D3FEND actions.
    Returns dict: action_name -> 0/1
    """
    action_df = pd.read_csv(action_per_ttps_path, index_col=0)
    action_names = action_df.columns.tolist()
    valid_ttps = [t for t in sample_ttps if t in action_df.index]

    if valid_ttps:
        action_sum = action_df.loc[valid_ttps].sum(axis=0)
        action_bin = (action_sum > 0).astype(int)
    else:
        action_bin = pd.Series(0, index=action_names)

    return action_bin.to_dict()


# ─────────────────────────────────────────────────────────────────────────────
# Compute action scores  [equivalent to Stage 7]
# score(a) = W * px(a) + beta * sum(sim_j * yj(a))
# ─────────────────────────────────────────────────────────────────────────────

def compute_action_scores(
    px: Dict[str, int],
    neighbors: List[Tuple[str, float]],
    hash_to_acts: Dict[str, Set[str]],
    all_actions: List[str],
    W: float,
    beta: float,
) -> List[Tuple[str, int, str, float]]:
    """
    Compute ranked action scores.
    Returns list of (sha256_placeholder, rank, action, score).
    """
    scores: Dict[str, float] = {}
    for action in all_actions:
        p = px.get(action, 0)
        neighbor_sum = 0.0
        for n_sha, n_sim in neighbors:
            yj = 1.0 if action in hash_to_acts.get(n_sha, set()) else 0.0
            neighbor_sum += n_sim * yj
        scores[action] = W * p + beta * neighbor_sum

    sorted_actions = sorted(all_actions, key=lambda a: scores[a], reverse=True)
    results = []
    for rank, action in enumerate(sorted_actions, 1):
        results.append((rank, action, scores[action]))
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Precision / Recall / F1 (same as auto_precision&recall.py)
# ─────────────────────────────────────────────────────────────────────────────

def extract_action_code(action_str):
    match = re.match(r"(D3-\w+)", action_str.strip())
    return match.group(1) if match else action_str.strip()


def load_ground_truth(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    gt = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        sha256, actions_str = row[0], row[1]
        if not sha256 or not actions_str:
            continue
        actions = set()
        for line in actions_str.strip().split("\n"):
            code = extract_action_code(line)
            if code:
                actions.add(code)
        if sha256 in gt:
            gt[sha256] = gt[sha256] | actions
        else:
            gt[sha256] = actions
    wb.close()
    return gt


def calculate_metrics_all_k(gt, scores_dict, top_k_levels):
    results = {}
    for sha256, predictions in scores_dict.items():
        if sha256 not in gt:
            continue
        gt_actions = gt[sha256]
        row = {}
        for k in top_k_levels:
            pred_actions = {p[1] for p in predictions[:k]}
            hits = pred_actions & gt_actions
            precision = len(hits) / len(pred_actions) if pred_actions else 0
            recall = len(hits) / len(gt_actions) if gt_actions else 0
            f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0
            row[f"precision_top{k}"] = round(precision, 4)
            row[f"recall_top{k}"] = round(recall, 4)
            row[f"f1_top{k}"] = round(f1, 4)
        results[sha256] = row
    return results


def unique_output_path(base_dir: Path, base_name: str, ext: str = ".csv") -> Path:
    candidate = base_dir / f"{base_name}{ext}"
    if not candidate.exists():
        return candidate
    i = 1
    while True:
        candidate = base_dir / f"{base_name}_{i}{ext}"
        if not candidate.exists():
            return candidate
        i += 1


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    setup_logging()
    parser = argparse.ArgumentParser(
        description="Eval from pre-computed Top-K similarity files",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--sample-dir", default="", help="Directory containing sample JSON reports")
    parser.add_argument("--topk-dir", default="", help="Directory containing top-k similarity CSV/XLSX files")
    parser.add_argument("--W", type=float, default=None, help="Self-weight W")
    parser.add_argument("--beta", type=float, default=None, help="Neighbor weight beta")
    parser.add_argument("--top-k", type=int, default=None, help="Use only top-k neighbors (must be <= available neighbors)")
    args = parser.parse_args()

    # ── Load config ──
    cfg = load_config()
    if cfg:
        print(f"[config] Loaded {CONFIG_FILE.name}")

    # Apply config defaults (CLI overrides config)
    W = args.W if args.W is not None else cfg.get("W", 0.3)
    beta = args.beta if args.beta is not None else cfg.get("beta", 0.7)
    top_k = args.top_k if args.top_k is not None else cfg.get("top_k", None)

    # Resolve data paths
    action_ttps_path = Path(cfg.get("action_ttps", "data/action_space/action_per_ttps.csv"))
    action_report_path = Path(cfg.get("action_report", "data/top_k/actionPerReport.xlsx"))
    unique_actions_path = Path(cfg.get("unique_actions", "data/top_k/unique_actions.csv"))
    ttp_unique_path = Path(cfg.get("ttp_unique", "data/unique/ttp_unique.csv"))

    # Make relative paths absolute from script dir
    for name in ("action_ttps_path", "action_report_path", "unique_actions_path", "ttp_unique_path"):
        p = locals()[name]
        if not p.is_absolute():
            locals()[name] = SCRIPT_DIR / p
    action_ttps_path = SCRIPT_DIR / action_ttps_path if not action_ttps_path.is_absolute() else action_ttps_path
    action_report_path = SCRIPT_DIR / action_report_path if not action_report_path.is_absolute() else action_report_path
    unique_actions_path = SCRIPT_DIR / unique_actions_path if not unique_actions_path.is_absolute() else unique_actions_path
    ttp_unique_path = SCRIPT_DIR / ttp_unique_path if not ttp_unique_path.is_absolute() else ttp_unique_path

    # ── Select sample directory ──
    sample_dir = args.sample_dir
    if not sample_dir:
        print("[GUI] Select the folder containing sample JSON reports...")
        sample_dir = pick_directory("Select sample JSON folder")
        if not sample_dir:
            print("No folder selected. Exiting.")
            sys.exit(0)
    sample_dir = Path(sample_dir)
    if not sample_dir.is_dir():
        print(f"[ERROR] Not a valid directory: {sample_dir}")
        sys.exit(1)

    json_files = sorted(sample_dir.glob("*.json"))
    if not json_files:
        print(f"[ERROR] No .json files found in {sample_dir}")
        sys.exit(1)

    # ── Select top-k similarity directory ──
    topk_dir = args.topk_dir
    if not topk_dir:
        print("[GUI] Select the folder containing top-k similarity files (CSV/XLSX)...")
        topk_dir = pick_directory("Select top-k similarity folder")
        if not topk_dir:
            print("No folder selected. Exiting.")
            sys.exit(0)
    topk_dir = Path(topk_dir)
    if not topk_dir.is_dir():
        print(f"[ERROR] Not a valid directory: {topk_dir}")
        sys.exit(1)

    # Index similarity files by sha256 (filename without extension)
    sim_files: Dict[str, Path] = {}
    for f in topk_dir.iterdir():
        if f.suffix.lower() in (".csv", ".xlsx") and f.is_file():
            sim_files[f.stem] = f

    if not sim_files:
        print(f"[ERROR] No CSV/XLSX files found in {topk_dir}")
        sys.exit(1)

    # ── Check required data files ──
    required = {
        "Action per TTPs": action_ttps_path,
        "Action per Report": action_report_path,
        "Unique Actions": unique_actions_path,
        "TTP unique": ttp_unique_path,
    }
    missing = [(n, p) for n, p in required.items() if not p.exists()]
    if missing:
        print("\n[ERROR] Missing required data files:")
        for n, p in missing:
            print(f"  {n}: {p}")
        sys.exit(1)

    # ── Load shared data ──
    print("\n[Data] Loading shared data files...")
    hash_to_acts = load_action_report(action_report_path)
    print(f"  actionPerReport: {len(hash_to_acts)} entries")

    actions_df = pd.read_csv(unique_actions_path)
    all_actions = actions_df["Action"].tolist()
    print(f"  unique_actions: {len(all_actions)} actions")

    ttp_features = load_feature_list(ttp_unique_path)
    print(f"  TTP features: {len(ttp_features)}")

    # ── Print config ──
    print(f"\n{'='*60}")
    print(f" Eval from Pre-computed Top-K Similarity")
    print(f"{'='*60}")
    print(f" Samples    : {len(json_files)} JSON files in {sample_dir.name}")
    print(f" Top-K dir  : {topk_dir.name} ({len(sim_files)} files)")
    print(f" W={W}  beta={beta}  top_k={top_k if top_k else 'all'}")
    print(f"{'='*60}")

    # ── Process each sample ──
    # scores_dict: sha256 -> list of (rank, action_code, score) for P/R/F1
    all_action_scores: Dict[str, List[Tuple[int, str, float]]] = {}
    processed = 0
    skipped = []

    for i, json_path in enumerate(json_files, 1):
        print(f"\n{'─'*60}")
        print(f" [{i}/{len(json_files)}] {json_path.name}")
        print(f"{'─'*60}")

        # Extract sha256 from JSON
        try:
            with open(json_path, encoding="utf-8", errors="replace") as f:
                report = json.load(f)
            sha256 = extract_hash(report)
        except Exception as e:
            print(f"  [ERROR] Cannot read JSON: {e}")
            skipped.append(json_path.name)
            continue

        if sha256 == "UNKNOWN_HASH":
            print(f"  [WARN] Could not extract sha256, skipping.")
            skipped.append(json_path.name)
            continue

        print(f"  SHA256: {sha256[:20]}...")

        # Find matching similarity file
        if sha256 not in sim_files:
            print(f"  [WARN] No similarity file found for this hash, skipping.")
            skipped.append(json_path.name)
            continue

        # Read pre-computed top-k neighbors
        neighbors = read_similarity_file(sim_files[sha256])
        if not neighbors:
            print(f"  [WARN] Empty similarity file, skipping.")
            skipped.append(json_path.name)
            continue

        # Truncate to top_k if configured
        if top_k is not None and top_k < len(neighbors):
            neighbors = neighbors[:top_k]

        print(f"  Top-{len(neighbors)} neighbors:")
        for j, (n_sha, n_sim) in enumerate(neighbors):
            print(f"    {j+1}. {n_sha[:20]}... | sim={n_sim:.4f}")

        # Extract TTPs from sample -> map to D3FEND actions (px)
        sample_ttps = extract_ttps(report)
        print(f"  Sample TTPs: {len(sample_ttps)}")
        px = map_sample_actions(sample_ttps, action_ttps_path)

        # Compute action scores
        scored = compute_action_scores(px, neighbors, hash_to_acts, all_actions, W, beta)

        # Store for P/R/F1: (rank, action_code, score)
        all_action_scores[sha256] = [(r, extract_action_code(a), s) for r, a, s in scored]

        # Print top 10
        print(f"  Top-10 recommended actions:")
        for rank, action, score in scored[:10]:
            print(f"    {rank:2d}. {action:<60s} {score:.4f}")

        processed += 1

    print(f"\n{'='*60}")
    print(f" Processing Summary: {processed}/{len(json_files)} succeeded")
    if skipped:
        print(f" Skipped: {', '.join(skipped)}")
    print(f"{'='*60}")

    if not all_action_scores:
        print("\n[Eval] No samples processed. Nothing to evaluate.")
        return

    # ── Calculate Precision / Recall / F1 ──
    if not GROUND_TRUTH_PATH.exists():
        print(f"\n[WARN] Ground truth not found: {GROUND_TRUTH_PATH}")
        print("[Eval] Skipping P/R/F1 calculation.")
        return

    gt = load_ground_truth(GROUND_TRUTH_PATH)
    print(f"\n[Eval] Ground truth: {len(gt)} samples")

    score_hashes = set(all_action_scores.keys())
    gt_hashes = set(gt.keys())
    matched = score_hashes & gt_hashes

    print(f"  Matched with ground truth: {len(matched)}")
    if score_hashes - gt_hashes:
        print(f"  In scores but NOT in ground truth: {len(score_hashes - gt_hashes)}")
    if gt_hashes - score_hashes:
        print(f"  In ground truth but NOT in scores: {len(gt_hashes - score_hashes)}")

    if not matched:
        print("\n[Eval] No matching samples. Cannot calculate metrics.")
        return

    filtered_scores = {h: v for h, v in all_action_scores.items() if h in matched}

    top_k_levels = [1, 2, 3, 4, 5]
    results = calculate_metrics_all_k(gt, filtered_scores, top_k_levels)

    if not results:
        print("\n[Eval] No results to compute.")
        return

    # Print averages
    print(f"\n[Eval] Metrics ({len(results)} samples):")
    for k in top_k_levels:
        avg_prec = sum(r[f"precision_top{k}"] for r in results.values()) / len(results)
        avg_rec = sum(r[f"recall_top{k}"] for r in results.values()) / len(results)
        avg_f1 = sum(r[f"f1_top{k}"] for r in results.values()) / len(results)
        print(f"  [top_{k}] Avg Precision: {avg_prec:.4f}, Avg Recall: {avg_rec:.4f}, Avg F1: {avg_f1:.4f}")

    # ── Save results ──
    all_results = []
    for sha256, metrics in results.items():
        row = {"sha256": sha256}
        row.update(metrics)
        all_results.append(row)

    # AVERAGE row
    avg_row = {"sha256": "AVERAGE"}
    for k in top_k_levels:
        avg_row[f"precision_top{k}"] = round(
            sum(r[f"precision_top{k}"] for r in all_results) / len(all_results), 4
        )
        avg_row[f"recall_top{k}"] = round(
            sum(r[f"recall_top{k}"] for r in all_results) / len(all_results), 4
        )
        avg_row[f"f1_top{k}"] = round(
            sum(r[f"f1_top{k}"] for r in all_results) / len(all_results), 4
        )
    all_results.append(avg_row)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = unique_output_path(RESULTS_DIR, RESULTS_BASE)

    fieldnames = ["sha256"]
    for k in top_k_levels:
        fieldnames.append(f"precision_top{k}")
        fieldnames.append(f"recall_top{k}")
        fieldnames.append(f"f1_top{k}")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_results)

    print(f"\n[Eval] Results saved to: {output_path}")

    print(f"\n{'='*60}")
    print(f" Eval complete!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
